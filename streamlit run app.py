import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import io
import os

# --- Page Configuration ---
st.set_page_config(
    page_title="Weekly ATM Transaction Monitoring Report",
    page_icon="💳",
    layout="centered"
)

# --- Helper Functions ---
def prepare_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        std = wb[sheet_name]
        wb.remove(std)
    return wb.create_sheet(sheet_name)

def find_column_index(header, keyword):
    for i, h in enumerate(header):
        if h and keyword.lower() in str(h).lower():
            return i
    return -1

def to_float_safe(value):
    if value is None:
        return 0.0
    try:
        s = str(value).replace(",", "").replace("MMK", "").strip()
        return float(s)
    except Exception:
        return 0.0

def add_terminal_details_sheet(wb_main, terminal_aggregation):
    ws_terminal = prepare_sheet(wb_main, "Terminal Details")
    
    headers = ["S/N", "Terminal ID", "Location", "Total Transaction Count", "Total use card", "On us card", "Off us card", "Total Amount", "Fee", "SUM", "On us amount", "Off us amount"]
    ws_terminal.append(headers)
    
    terminal_info = [
        [1, "09950001", "Dagon (1)"], [2, "09950002", "Botahtaung"], [3, "09950003", "MIP"],
        [4, "09950004", "MDY"], [5, "09950005", "NPT MOC 40"], [6, "09950006", "Pathein"],
        [7, "09950007", "Naypyitaw Br"], [8, "09950008", "Mawlamyaing"], [9, "09950009", "Shwe Pauk Kan"],
        [10, "09950010", "Magway"], [11, "09950013", "Dagon (2)"], [12, "09950014", "Monywa"],
        [13, "09950015", "Hpa-an"], [14, "09950016", "Dawei"], [15, "09950017", "DSK"],
        [16, "09950018", "BGO"], [17, "09950019", "MKN"], [18, "09950021", "HTY"],
        [19, "09950024", "TGI"], [20, "09950025", "MICT Park"], [21, "09950026", "NPT MOC 11"],
        [22, "09950027", "NPT MOC 11 (2)"], [23, "09950028", "NPT MOC 40 (2)"], [24, "09950029", "NPT Hluttaw"]
    ]
    
    ygn_public_ids = ["09950003", "09950025"]
    ygn_branch_ids = ["09950001", "09950002", "09950009", "09950013", "09950017", "09950021"]
    other_public_ids = ["09950005", "09950026", "09950027", "09950028", "09950029"]
    other_branch_ids = ["09950004", "09950006", "09950007", "09950008", "09950010", "09950014", "09950015", "09950016", "09950018", "09950019", "09950024"]

    grand_totals = {
        'count': 0, 'use_card': 0, 'on_us_card': 0, 'off_us_card': 0,
        'amount': 0.0, 'fee': 0.0, 'sum': 0.0, 'on_us_amt': 0.0, 'off_us_amt': 0.0
    }

    ygn_public_sum, ygn_branch_sum, other_public_sum, other_branch_sum = 0.0, 0.0, 0.0, 0.0
    
    for info in terminal_info:
        sn, terminal_id, location = info
        terminal_data = terminal_aggregation.get(terminal_id, {
            'total_transaction_count': 0, 'total_transaction_amount': 0.0, 'total_fee': 0.0,
            'on_us_count': 0, 'off_us_count': 0, 'on_us_amount': 0.0, 'off_us_amount': 0.0,
            'unique_card_count': 0, 'card_numbers': set()
        })
        
        transaction_count = terminal_data['total_transaction_count']
        total_amount = terminal_data['total_transaction_amount']
        total_fee = terminal_data['total_fee']
        sum_value = total_amount - total_fee
        
        on_us_count = terminal_data['on_us_count']
        off_us_count = terminal_data['off_us_count']
        total_use_card = terminal_data['unique_card_count']
        on_us_amount = terminal_data['on_us_amount']
        off_us_amount = terminal_data['off_us_amount']
        
        grand_totals['count'] += transaction_count
        grand_totals['use_card'] += total_use_card
        grand_totals['on_us_card'] += on_us_count
        grand_totals['off_us_card'] += off_us_count
        grand_totals['amount'] += total_amount
        grand_totals['fee'] += total_fee
        grand_totals['sum'] += sum_value
        grand_totals['on_us_amt'] += on_us_amount
        grand_totals['off_us_amt'] += off_us_amount

        if terminal_id in ygn_public_ids:
            ygn_public_sum += sum_value
        elif terminal_id in ygn_branch_ids:
            ygn_branch_sum += sum_value
        elif terminal_id in other_public_ids:
            other_public_sum += sum_value
        elif terminal_id in other_branch_ids:
            other_branch_sum += sum_value
            
        ws_terminal.append([
            sn, terminal_id, location, transaction_count, total_use_card, on_us_count, off_us_count,
            total_amount, total_fee, sum_value, on_us_amount, off_us_amount
        ])
    
    grand_total_row_idx = ws_terminal.max_row + 1
    ws_terminal.append([
        "", "", "Total SUM", 
        grand_totals['count'], grand_totals['use_card'], grand_totals['on_us_card'], grand_totals['off_us_card'],
        grand_totals['amount'], grand_totals['fee'], grand_totals['sum'], grand_totals['on_us_amt'], grand_totals['off_us_amt']
    ])

    ws_terminal.append([])
    
    summary_start_row = ws_terminal.max_row + 1
    ws_terminal.append(["", "", "YGN Public Total SUM", "", "", "", "", "", "", ygn_public_sum, "", ""])
    ws_terminal.append(["", "", "YGN Branch Total SUM", "", "", "", "", "", "", ygn_branch_sum, "", ""])
    ws_terminal.append(["", "", "Other Public Total SUM", "", "", "", "", "", "", other_public_sum, "", ""])
    ws_terminal.append(["", "", "Other Branch Total SUM", "", "", "", "", "", "", other_branch_sum, "", ""])
    summary_end_row = ws_terminal.max_row

    for row in range(2, ws_terminal.max_row + 1):
        if row < grand_total_row_idx:
            for col in [4, 5, 6, 7]:
                ws_terminal.cell(row=row, column=col).number_format = '#,##0'
            for col in [8, 9, 10, 11, 12]:
                ws_terminal.cell(row=row, column=col).number_format = '#,##0.00'
        elif row == grand_total_row_idx:
            for col in range(1, 13):
                cell = ws_terminal.cell(row=row, column=col)
                cell.font = Font(bold=True)
                if col in [4, 5, 6, 7]:
                    cell.number_format = '#,##0'
                elif col in [8, 9, 10, 11, 12]:
                    cell.number_format = '#,##0.00'
        elif summary_start_row <= row <= summary_end_row:
            ws_terminal.cell(row=row, column=3).font = Font(bold=True)
            sum_cell = ws_terminal.cell(row=row, column=10)
            sum_cell.number_format = '#,##0.00'
            sum_cell.font = Font(bold=True)
            
    for column in ws_terminal.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_terminal.column_dimensions[column_letter].width = (max_length + 2)

def add_terminal_transaction_sheets(wb_main, terminal_transactions, header):
    for terminal_id, transactions in terminal_transactions.items():
        if not transactions:  
            continue
        sheet_name = f"T_{terminal_id}"
        ws = prepare_sheet(wb_main, sheet_name)
        ws.append(header)
        
        for row in transactions:
            ws.append([cell.value for cell in row])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min((max_length + 2), 30)

def process_transaction_bytes(file_bytes, file_name):
    # Streamlit file upload gives bytes. Read into openpyxl using BytesIO
    ext = os.path.splitext(file_name)[1].lower()
    
    if ext == ".xls":
        # Convert legacy xls to dataframe via xlrd, then create an openpyxl memory stream
        xls_df = pd.read_excel(io.BytesIO(file_bytes), engine='xlrd')
        wb_data = Workbook()
        ws_data = wb_data.active
        ws_data.title = "Sheet1"
        ws_data.append(list(xls_df.columns))
        for r in xls_df.values.tolist():
            ws_data.append(r)
    else:
        wb_data = load_workbook(io.BytesIO(file_bytes))

    ws_data = wb_data["Sheet1"] if "Sheet1" in wb_data.sheetnames else wb_data.active

    wb_main = Workbook()
    ws_main = wb_main.active
    ws_main.title = "Main"

    for row in ws_data.iter_rows():
        ws_main.append([cell.value for cell in row])
    wb_data.close()

    header = None
    header_row = 1
    for row_num in range(1, min(4, ws_main.max_row + 1)):
        test_header = [cell.value for cell in ws_main[row_num]]
        header_text = " ".join([str(h).lower() for h in test_header if h])
        if any(keyword in header_text for keyword in ["terminal", "amount", "fee", "ref"]):
            header = test_header
            header_row = row_num
            break
    
    if header is None:
        header = [cell.value for cell in ws_main[1]]
    
    amount_idx = find_column_index(header, "amount")
    fee_idx = find_column_index(header, "fee")
    card_idx = find_column_index(header, "card")
    refno_idx = find_column_index(header, "ref")
    
    terminal_id_idx = -1
    for keyword in ["terminal id", "terminal", "tid"]:
        terminal_id_idx = find_column_index(header, keyword)
        if terminal_id_idx != -1:
            break

    status_idx = find_column_index(header, "status")
    tranx_amount_idx = find_column_index(header, "tranx amount")
    if tranx_amount_idx == -1:
        tranx_amount_idx = find_column_index(header, "transaction amount")
    if tranx_amount_idx == -1:
        tranx_amount_idx = amount_idx  
    
    if amount_idx == -1 or fee_idx == -1 or refno_idx == -1 or terminal_id_idx == -1:
        st.error("❌ Required columns missing in Excel mapping hierarchy.")
        return None

    terminal_ids = [
        "09950001", "09950002", "09950003", "09950004", "09950005", "09950006", "09950007", "09950008", "09950009", "09950010",
        "09950013", "09950014", "09950015", "09950016", "09950017", "09950018", "09950019", "09950021", "09950024", "09950025",
        "09950026", "09950027", "09950028", "09950029"
    ]
    
    terminal_aggregation = {tid: {
        'total_transaction_count': 0, 'total_transaction_amount': 0.0, 'total_fee': 0.0,
        'on_us_count': 0, 'off_us_count': 0, 'on_us_amount': 0.0, 'off_us_amount': 0.0,
        'unique_card_count': 0, 'card_numbers': set()
    } for tid in terminal_ids}
    
    terminal_transactions = {tid: [] for tid in terminal_ids}

    reversal_ref_nos = set()
    for row in ws_main.iter_rows(min_row=header_row + 1):
        txt = " ".join([str(cell.value).strip() if cell.value else '' for cell in row])
        ref_no = str(row[refno_idx].value).strip() if row[refno_idx].value else ""
        if "reversal" in txt.lower() and ref_no:
            reversal_ref_nos.add(ref_no)

    rows_to_delete = []
    for row_idx in range(2, ws_main.max_row + 1):  
        row = ws_main[row_idx]
        if status_idx != -1 and row[status_idx].value is not None:
            if str(row[status_idx].value).strip() == "9":
                rows_to_delete.append(row_idx)
    
    for row_idx in sorted(rows_to_delete, reverse=True):
        ws_main.delete_rows(row_idx)
    
    for row_idx in range(2, ws_main.max_row + 1):  
        row = ws_main[row_idx]
        ref_no = str(row[refno_idx].value).strip() if row[refno_idx].value else ""
        if ref_no in reversal_ref_nos:
            continue

        terminal_id = str(row[terminal_id_idx].value).strip() if terminal_id_idx != -1 and row[terminal_id_idx].value else ""
        amount = to_float_safe(row[amount_idx].value)
        fee = to_float_safe(row[fee_idx].value)
        card_number = str(row[card_idx].value).strip() if card_idx != -1 and row[card_idx].value else ""
        
        if terminal_id in terminal_ids:
            terminal_transactions[terminal_id].append(row)
            terminal_aggregation[terminal_id]['total_transaction_amount'] += amount
            terminal_aggregation[terminal_id]['total_fee'] += fee
            terminal_aggregation[terminal_id]['total_transaction_count'] += 1
            
            if card_number and card_number.startswith("950316"):
                terminal_aggregation[terminal_id]['on_us_count'] += 1
                terminal_aggregation[terminal_id]['on_us_amount'] += amount
            else:
                terminal_aggregation[terminal_id]['off_us_count'] += 1
                terminal_aggregation[terminal_id]['off_us_amount'] += amount
            
            if card_number:
                terminal_aggregation[terminal_id]['card_numbers'].add(card_number)
                terminal_aggregation[terminal_id]['unique_card_count'] = len(terminal_aggregation[terminal_id]['card_numbers'])
        
        if status_idx != -1 and row[status_idx].value is not None:
            row[status_idx].value = str(row[status_idx].value).replace('9', '')
        
        if tranx_amount_idx != -1 and row[tranx_amount_idx].value is not None:
            row[tranx_amount_idx].value = str(row[tranx_amount_idx].value).replace('MMK', '')

    add_terminal_transaction_sheets(wb_main, terminal_transactions, header)
    add_terminal_details_sheet(wb_main, terminal_aggregation)
    
    # Save output to buffer memory for down-stream web downloads
    out_buffer = io.BytesIO()
    wb_main.save(out_buffer)
    out_buffer.seek(0)
    return out_buffer

# --- Web UI Structure ---
st.title("💳 ATM Transaction Processor")
st.subheader("Weekly & Monthly ATM Transaction Monitoring Report")
st.write("Upload your Excel spreadsheet (`.xls`, `.xlsx`, `.xlsm`) generated from FeelSwitch below:")

uploaded_file = st.file_uploader("Choose an Excel file", type=["xls", "xlsx", "xlsm"])

if uploaded_file is not None:
    st.info("🔄 Processing transaction dataset pipeline...")
    
    # Run the transaction processing pipeline
    processed_file = process_transaction_bytes(uploaded_file.getvalue(), uploaded_file.name)
    
    if processed_file:
        st.success("✅ Log tracking successfully summarized and built!")
        
        # UI Web File downloader button triggering native browser action
        st.download_button(
            label="📥 Download Processed Report",
            data=processed_file,
            file_name="Processed_ATM_Monitoring_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# --- Web Footer ---
st.markdown("---")
st.caption("© Version 2 (Web Variant) | Developed by Bo Bo Tun")